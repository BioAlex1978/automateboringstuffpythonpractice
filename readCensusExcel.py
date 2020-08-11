#! python3
# readCensusExcel.py - Tabulates population and number of census tracts for each county
# Automate the Boring Stuff with Python 2e, chapter 13, pg 308

import openpyxl, pprint

print('Opening workbook...')
wb = openpyxl.load_workbook('censuspopdata.xlsx')
# the book has us use the actual sheetname string 'Population by Census Tract', but that seems typo prone
# and this accomplishes exactly the same thing
sheet = wb[wb.sheetnames[0]]
countyData = {} #empty dictionary to hold county data, w/ States as first level keys which map to dictionaries of dictionaries
# the data structure of countyData will be countyData[state abbrev][county]['tracts'/'pop']

# TODO: Fill in countyData with each county's population and tracts
print('Reading rows...')
# Iterate through the sheet, starting with row 2, because row 1 is column titles
for row in range(2, sheet.max_row + 1):
    # Each row in the spreadsheet has data for one census tract.
    # Column A is tract numbers, and we don't need that for this, so we start with column B
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value
    
    # Make sure the key for this state exists
    countyData.setdefault(state, {})
    # Make sure the key for this county in this state exists
    countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})
    
    # Each row represents one census tract, so increment by one
    countyData[state][county]['tracts'] += 1
    # Increase the county pop by the pop in this row
    countyData[state][county]['pop'] += int(pop)
    
# Open a new text file and write the contents of countyData to it.
# In this case, we are compiling it into a variable called allData in another python file
# which we can then import into other scripts if we want to use it
# pprint.pformat() it converts the dictionary to a huge text string for adding to the 'allData = ' string
# that we are writing to the new file
print('Writing results...')
resultFile = open('census2010.py', 'w')
resultFile.write('allData = ' + pprint.pformat(countyData))
resultFile.close()
print('Done.')

# (Optional challenge for myself): Create a new spreadsheet file and write the contents of countyData to that
# This file will have 2 sheets: one w/ county data and one w/ state totals
print('Creating a new spreadsheet with the results...')
cdWorkbook = openpyxl.Workbook() # create the new workbook
# Set up the sheets
cdSheet1 = cdWorkbook.active
cdSheet1.title = '2010 Census County & State Data'
colhead1 = ['State', 'County', 'Population', 'Tracts'] # Column headers to go in the A row for the first sheet
for col in colhead1:
    cdSheet1.cell(row = 1, column = (colhead1.index(col) + 1)).value = col
cdWorkbook.create_sheet(index=1, title='2010 Census State Totals')
cdSheet2 = cdWorkbook[cdWorkbook.sheetnames[1]]
colhead2 = ['State', 'Population', 'Tracts'] # Column headers to go in the A row for the second sheet
for col in colhead2:
    cdSheet2.cell(row = 1, column = (colhead2.index(col) + 1)).value = col
    
# Now that row one is set up in both sheets, start adding data from row 2
sheet1row = 2
sheet2row = 2

# Iterate through countyData, and write the info to the appropriate cells and sheets
# Iterate through the states
for state in countyData:
    totalPop = 0 # set the starting total population value to zero
    totalTracts = 0 # set the starting total tracts to zero
    # Iterate through the counties
    for county in countyData[state]:
        # write the info into the current row in sheet 1
        cdSheet1['A' + str(sheet1row)] = state
        cdSheet1['B' + str(sheet1row)] = county
        cdSheet1['C' + str(sheet1row)] = countyData[state][county]['pop']
        cdSheet1['D' + str(sheet1row)] = countyData[state][county]['tracts']
        # update the state totals
        totalPop += countyData[state][county]['pop']
        totalTracts += countyData[state][county]['tracts']
        # increment the row number
        sheet1row += 1
    # Write the state data into the current row in sheet 2
    cdSheet2['A' + str(sheet2row)] = state
    cdSheet2['B' + str(sheet2row)] = totalPop
    cdSheet2['C' + str(sheet2row)] = totalTracts
    # increment the row number
    sheet2row += 1

# save the workbook
cdWorkbook.save('CensusCountyStateData.xlsx')
print('Spreadsheet saved as "CensusCountyStateData.xlsx".')